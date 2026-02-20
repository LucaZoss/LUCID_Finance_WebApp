"""
FastAPI backend for LUCID Finance WebApp.

Provides REST API endpoints for:
- Transaction management (CRUD + upload)
- Budget planning
- Dashboard summaries
"""

import os
import shutil
from datetime import date, datetime
from decimal import Decimal
from pathlib import Path
from typing import List, Optional
from io import BytesIO

from dotenv import load_dotenv
from fastapi import Depends, FastAPI, File, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

from ..data_pipeline.config import PipelineConfig, DatabaseConfig
from ..data_pipeline.models import DatabaseManager, Transaction, BudgetPlan, Category, User, CategorizationRule
from ..data_pipeline.pipeline import TransactionPipeline
from .auth import (
    authenticate_user,
    create_access_token,
    get_current_user,
    get_admin_user,
    get_password_hash,
)

# Initialize app
app = FastAPI(
    title="LUCID Finance API",
    description="Personal budgeting application API",
    version="0.1.0",
)

# CORS middleware for React frontend
app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:5173", "http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Database setup
db_config = DatabaseConfig()
db_manager = DatabaseManager(db_config)
pipeline_config = PipelineConfig()


# Pydantic models for API
class TransactionResponse(BaseModel):
    id: int
    date: date
    type: str
    category: str
    amount: float
    description: Optional[str]
    source: str
    month: int
    year: int
    source_file: Optional[str]

    class Config:
        from_attributes = True


class TransactionUpdate(BaseModel):
    type: Optional[str] = None
    category: Optional[str] = None


class BudgetPlanResponse(BaseModel):
    id: int
    type: str
    category: str
    year: int
    month: Optional[int]
    amount: float

    class Config:
        from_attributes = True


class BudgetPlanCreate(BaseModel):
    type: str
    category: str
    year: int
    month: Optional[int] = None
    amount: float


class CategoryInfo(BaseModel):
    type: str
    categories: List[str]


class CategoryResponse(BaseModel):
    id: int
    name: str
    type: str
    is_active: bool
    display_order: int
    created_at: datetime

    class Config:
        from_attributes = True


class CategoryCreate(BaseModel):
    name: str
    type: str
    display_order: int = 0


class CategoryUpdate(BaseModel):
    name: Optional[str] = None
    type: Optional[str] = None
    is_active: Optional[bool] = None
    display_order: Optional[int] = None


class SummaryItem(BaseModel):
    type: str
    category: str
    budget: float
    actual: float
    remaining: float
    percent_complete: float


class DashboardSummary(BaseModel):
    year: int
    month: Optional[int]
    income: List[SummaryItem]
    expenses: List[SummaryItem]
    savings: List[SummaryItem]
    totals: dict
    fixed_cost_ratio: float
    previous_period: dict


class LoginRequest(BaseModel):
    username: str
    password: str


class LoginResponse(BaseModel):
    access_token: str
    token_type: str
    username: str
    is_admin: bool


class UserCreate(BaseModel):
    username: str
    password: str
    full_name: Optional[str] = None
    is_admin: bool = False


class UserResponse(BaseModel):
    id: int
    username: str
    full_name: Optional[str]
    is_active: bool
    is_admin: bool
    created_at: datetime

    class Config:
        from_attributes = True


class RuleCreate(BaseModel):
    pattern: str
    case_sensitive: bool = False
    amount_operator: Optional[str] = None  # eq, gte, lte, gt, lt
    amount_value: Optional[float] = None
    type: str
    category: str
    priority: int = 0


class RuleUpdate(BaseModel):
    pattern: Optional[str] = None
    case_sensitive: Optional[bool] = None
    amount_operator: Optional[str] = None
    amount_value: Optional[float] = None
    type: Optional[str] = None
    category: Optional[str] = None
    priority: Optional[int] = None
    is_active: Optional[bool] = None


class RuleResponse(BaseModel):
    id: int
    pattern: str
    case_sensitive: bool
    amount_operator: Optional[str]
    amount_value: Optional[float]
    type: str
    category: str
    priority: int
    is_active: bool
    user_id: Optional[int]
    created_at: datetime
    updated_at: datetime

    class Config:
        from_attributes = True


# ============== Authentication Endpoints ==============

@app.post("/api/auth/login", response_model=LoginResponse)
def login(credentials: LoginRequest):
    """Authenticate user and return JWT token."""
    user = authenticate_user(credentials.username, credentials.password)
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Incorrect username or password"
        )

    access_token = create_access_token(
        data={"sub": user["username"], "user_id": user["id"], "is_admin": user["is_admin"]}
    )

    return LoginResponse(
        access_token=access_token,
        token_type="bearer",
        username=user["username"],
        is_admin=user["is_admin"]
    )


@app.post("/api/auth/users", response_model=UserResponse)
def create_user(user_data: UserCreate, current_user: dict = Depends(get_admin_user)):
    """Create a new user (admin only)."""
    session = db_manager.get_session()
    try:
        # Check if username already exists
        existing = session.query(User).filter(User.username == user_data.username).first()
        if existing:
            raise HTTPException(status_code=400, detail="Username already exists")

        # Create new user
        new_user = User(
            username=user_data.username,
            hashed_password=get_password_hash(user_data.password),
            full_name=user_data.full_name,
            is_admin=user_data.is_admin,
            is_active=True
        )
        session.add(new_user)
        session.commit()
        session.refresh(new_user)

        return UserResponse.model_validate(new_user)
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.get("/api/auth/users", response_model=List[UserResponse])
def list_users(current_user: dict = Depends(get_admin_user)):
    """List all users (admin only)."""
    session = db_manager.get_session()
    try:
        users = session.query(User).all()
        return [UserResponse.model_validate(u) for u in users]
    finally:
        session.close()


@app.get("/api/auth/me")
def get_current_user_info(current_user: dict = Depends(get_current_user)):
    """Get current authenticated user info."""
    return current_user


# ============== Categorization Rules Endpoints ==============

@app.get("/api/rules", response_model=List[RuleResponse])
def get_rules(
    is_active: Optional[bool] = None,
    current_user: dict = Depends(get_current_user),
):
    """Get all categorization rules, ordered by priority (highest first)."""
    session = db_manager.get_session()
    try:
        query = session.query(CategorizationRule).order_by(
            CategorizationRule.priority.desc(),
            CategorizationRule.created_at.desc()
        )

        if is_active is not None:
            query = query.filter(CategorizationRule.is_active == is_active)

        rules = query.all()
        return [RuleResponse.model_validate(rule) for rule in rules]
    finally:
        session.close()


@app.post("/api/rules", response_model=RuleResponse)
def create_rule(
    rule_data: RuleCreate,
    current_user: dict = Depends(get_current_user),
):
    """Create a new categorization rule."""
    session = db_manager.get_session()
    try:
        new_rule = CategorizationRule(
            pattern=rule_data.pattern,
            case_sensitive=rule_data.case_sensitive,
            amount_operator=rule_data.amount_operator,
            amount_value=rule_data.amount_value,
            type=rule_data.type,
            category=rule_data.category,
            priority=rule_data.priority,
            is_active=True,
            user_id=None  # For now, all rules are global
        )
        session.add(new_rule)
        session.commit()
        session.refresh(new_rule)

        return RuleResponse.model_validate(new_rule)
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.patch("/api/rules/{rule_id}", response_model=RuleResponse)
def update_rule(
    rule_id: int,
    rule_data: RuleUpdate,
    current_user: dict = Depends(get_current_user),
):
    """Update a categorization rule."""
    session = db_manager.get_session()
    try:
        rule = session.query(CategorizationRule).filter(CategorizationRule.id == rule_id).first()
        if not rule:
            raise HTTPException(status_code=404, detail="Rule not found")

        # Update fields if provided
        if rule_data.pattern is not None:
            rule.pattern = rule_data.pattern
        if rule_data.case_sensitive is not None:
            rule.case_sensitive = rule_data.case_sensitive
        if rule_data.amount_operator is not None:
            rule.amount_operator = rule_data.amount_operator
        if rule_data.amount_value is not None:
            rule.amount_value = rule_data.amount_value
        if rule_data.type is not None:
            rule.type = rule_data.type
        if rule_data.category is not None:
            rule.category = rule_data.category
        if rule_data.priority is not None:
            rule.priority = rule_data.priority
        if rule_data.is_active is not None:
            rule.is_active = rule_data.is_active

        session.commit()
        session.refresh(rule)

        return RuleResponse.model_validate(rule)
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.delete("/api/rules/{rule_id}")
def delete_rule(
    rule_id: int,
    current_user: dict = Depends(get_current_user),
):
    """Delete a categorization rule."""
    session = db_manager.get_session()
    try:
        rule = session.query(CategorizationRule).filter(CategorizationRule.id == rule_id).first()
        if not rule:
            raise HTTPException(status_code=404, detail="Rule not found")

        session.delete(rule)
        session.commit()

        return {"message": "Rule deleted successfully", "id": rule_id}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.post("/api/rules/apply")
def apply_rules_to_transactions(
    current_user: dict = Depends(get_current_user),
):
    """
    Re-categorize existing transactions based on current active rules.
    This checks all transactions and updates their type/category if they match a rule.
    """
    from ..data_pipeline.transformers import TransactionTransformer

    session = db_manager.get_session()
    try:
        # Get all active rules
        rules = session.query(CategorizationRule).filter(
            CategorizationRule.is_active.is_(True)
        ).order_by(
            CategorizationRule.priority.desc(),
            CategorizationRule.created_at.desc()
        ).all()

        if not rules:
            return {"message": "No active rules to apply", "updated_count": 0}

        # Get all transactions for current user
        transactions = session.query(Transaction).filter(
            Transaction.user_id == current_user["id"]
        ).all()

        if not transactions:
            return {"message": "No transactions to process", "updated_count": 0}

        # Create transformer to check rules
        transformer = TransactionTransformer(pipeline_config, db_manager)

        updated_count = 0

        for transaction in transactions:
            # Check if transaction matches any rule
            description = transaction.description or ""
            amount = float(transaction.amount)

            # Use transformer's rule checking logic
            match = transformer._check_custom_rules(description, amount)

            if match:
                new_type, new_category = match
                # Only update if different
                if transaction.type != new_type or transaction.category != new_category:
                    transaction.type = new_type
                    transaction.category = new_category
                    updated_count += 1

        session.commit()

        return {
            "message": f"Successfully re-categorized {updated_count} transactions",
            "updated_count": updated_count,
            "total_transactions": len(transactions),
            "active_rules": len(rules)
        }

    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


# ============== Transaction Endpoints ==============

@app.get("/api/transactions", response_model=List[TransactionResponse])
def get_transactions(
    year: Optional[int] = None,
    month: Optional[int] = None,
    type: Optional[str] = None,
    category: Optional[str] = None,
    amount_min: Optional[float] = None,
    amount_max: Optional[float] = None,
    limit: int = Query(default=500, le=5000),
    offset: int = 0,
    current_user: dict = Depends(get_current_user),
):
    """Get transactions with optional filters."""
    session = db_manager.get_session()
    try:
        # Filter by current user
        query = session.query(Transaction).filter(Transaction.user_id == current_user["id"])

        if year:
            query = query.filter(Transaction.year == year)
        if month:
            query = query.filter(Transaction.month == month)
        if type:
            query = query.filter(Transaction.type == type)
        if category:
            query = query.filter(Transaction.category == category)
        if amount_min is not None:
            query = query.filter(Transaction.amount >= amount_min)
        if amount_max is not None:
            query = query.filter(Transaction.amount <= amount_max)

        query = query.order_by(Transaction.date.desc())
        transactions = query.offset(offset).limit(limit).all()

        return [TransactionResponse.model_validate(t) for t in transactions]
    finally:
        session.close()


@app.get("/api/transactions/{transaction_id}", response_model=TransactionResponse)
def get_transaction(transaction_id: int, current_user: dict = Depends(get_current_user)):
    """Get a single transaction by ID."""
    session = db_manager.get_session()
    try:
        transaction = session.query(Transaction).filter(
            Transaction.id == transaction_id,
            Transaction.user_id == current_user["id"]
        ).first()
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")
        return TransactionResponse.model_validate(transaction)
    finally:
        session.close()


@app.patch("/api/transactions/{transaction_id}", response_model=TransactionResponse)
def update_transaction(transaction_id: int, update: TransactionUpdate, current_user: dict = Depends(get_current_user)):
    """Update a transaction's type and/or category."""
    session = db_manager.get_session()
    try:
        transaction = session.query(Transaction).filter(
            Transaction.id == transaction_id,
            Transaction.user_id == current_user["id"]
        ).first()
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")

        if update.type is not None:
            transaction.type = update.type
        if update.category is not None:
            transaction.category = update.category

        session.commit()
        session.refresh(transaction)
        return TransactionResponse.model_validate(transaction)
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.delete("/api/transactions/{transaction_id}")
def delete_transaction(transaction_id: int, current_user: dict = Depends(get_current_user)):
    """Delete a transaction."""
    session = db_manager.get_session()
    try:
        transaction = session.query(Transaction).filter(
            Transaction.id == transaction_id,
            Transaction.user_id == current_user["id"]
        ).first()
        if not transaction:
            raise HTTPException(status_code=404, detail="Transaction not found")

        session.delete(transaction)
        session.commit()
        return {"message": "Transaction deleted"}
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.post("/api/transactions/bulk-update")
def bulk_update_transactions(updates: dict, current_user: dict = Depends(get_current_user)):
    """Bulk update transactions by criteria (e.g., reclassify all matching transactions)."""
    session = db_manager.get_session()
    try:
        # Example: {"description_contains": "epargne", "updates": {"type": "Savings", "category": "Rent Guarantee"}}
        description_filter = updates.get("description_contains", "").lower()
        category_filter = updates.get("category_filter")
        new_type = updates.get("updates", {}).get("type")
        new_category = updates.get("updates", {}).get("category")

        if not description_filter and not category_filter:
            raise HTTPException(status_code=400, detail="Must provide filter criteria")
        if not new_type and not new_category:
            raise HTTPException(status_code=400, detail="Must provide updates")

        # Filter by current user
        query = session.query(Transaction).filter(Transaction.user_id == current_user["id"])
        if description_filter:
            query = query.filter(Transaction.description.ilike(f"%{description_filter}%"))
        if category_filter:
            query = query.filter(Transaction.category == category_filter)

        transactions = query.all()
        count = 0
        for trans in transactions:
            if new_type:
                trans.type = new_type
            if new_category:
                trans.category = new_category
            count += 1

        session.commit()
        return {"message": f"Updated {count} transactions"}
    except HTTPException:
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


# ============== Upload Endpoints ==============

@app.post("/api/upload")
async def upload_csv(
    ubs_file: Optional[UploadFile] = File(None),
    cc_file: Optional[UploadFile] = File(None),
):
    """Upload UBS and/or CC CSV files for processing."""
    if not ubs_file and not cc_file:
        raise HTTPException(status_code=400, detail="At least one file must be provided")

    # Create temp directory for uploads
    upload_dir = Path("temp_uploads")
    upload_dir.mkdir(exist_ok=True)

    ubs_path = None
    cc_path = None

    try:
        # Save uploaded files
        if ubs_file:
            ubs_path = upload_dir / ubs_file.filename
            with open(ubs_path, "wb") as f:
                shutil.copyfileobj(ubs_file.file, f)

        if cc_file:
            cc_path = upload_dir / cc_file.filename
            with open(cc_path, "wb") as f:
                shutil.copyfileobj(cc_file.file, f)

        # Process files
        pipeline = TransactionPipeline()
        stats = pipeline.process_files(
            ubs_file=str(ubs_path) if ubs_path else None,
            cc_file=str(cc_path) if cc_path else None,
            force=False,
        )

        return {
            "message": "Files processed successfully",
            "stats": stats,
        }

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

    finally:
        # Cleanup temp files
        if ubs_path and ubs_path.exists():
            ubs_path.unlink()
        if cc_path and cc_path.exists():
            cc_path.unlink()


# ============== Budget Endpoints ==============

@app.get("/api/budgets", response_model=List[BudgetPlanResponse])
def get_budgets(year: Optional[int] = None, current_user: dict = Depends(get_current_user)):
    """Get all budget plans, optionally filtered by year."""
    session = db_manager.get_session()
    try:
        query = session.query(BudgetPlan).filter(BudgetPlan.user_id == current_user["id"])
        if year:
            query = query.filter(BudgetPlan.year == year)
        budgets = query.all()
        return [BudgetPlanResponse.model_validate(b) for b in budgets]
    finally:
        session.close()


@app.post("/api/budgets", response_model=BudgetPlanResponse)
def create_budget(budget: BudgetPlanCreate, auto_populate: bool = True, current_user: dict = Depends(get_current_user)):
    """Create or update a budget plan with optional auto-population."""
    session = db_manager.get_session()
    try:
        # Check if budget already exists for current user
        existing = session.query(BudgetPlan).filter(
            BudgetPlan.user_id == current_user["id"],
            BudgetPlan.type == budget.type,
            BudgetPlan.category == budget.category,
            BudgetPlan.year == budget.year,
            BudgetPlan.month == budget.month,
        ).first()

        if existing:
            existing.amount = Decimal(str(budget.amount))
        else:
            existing = BudgetPlan(
                user_id=current_user["id"],
                type=budget.type,
                category=budget.category,
                year=budget.year,
                month=budget.month,
                amount=Decimal(str(budget.amount)),
            )
            session.add(existing)

        session.commit()
        session.refresh(existing)

        # Auto-populate monthly budgets from yearly (or vice versa)
        if auto_populate:
            if budget.month is None:
                # Yearly budget entered → create/update 12 monthly budgets
                monthly_amount = Decimal(str(budget.amount)) / 12
                for month_num in range(1, 13):
                    monthly_budget = session.query(BudgetPlan).filter(
                        BudgetPlan.user_id == current_user["id"],
                        BudgetPlan.type == budget.type,
                        BudgetPlan.category == budget.category,
                        BudgetPlan.year == budget.year,
                        BudgetPlan.month == month_num,
                    ).first()

                    if monthly_budget:
                        monthly_budget.amount = monthly_amount
                    else:
                        monthly_budget = BudgetPlan(
                            user_id=current_user["id"],
                            type=budget.type,
                            category=budget.category,
                            year=budget.year,
                            month=month_num,
                            amount=monthly_amount,
                        )
                        session.add(monthly_budget)
            else:
                # Monthly budget entered → update yearly budget (sum all months)
                all_monthly = session.query(BudgetPlan).filter(
                    BudgetPlan.user_id == current_user["id"],
                    BudgetPlan.type == budget.type,
                    BudgetPlan.category == budget.category,
                    BudgetPlan.year == budget.year,
                    BudgetPlan.month.isnot(None),
                ).all()

                if len(all_monthly) == 12:
                    # All 12 months exist, calculate yearly total
                    yearly_total = sum(Decimal(str(b.amount)) for b in all_monthly)
                    yearly_budget = session.query(BudgetPlan).filter(
                        BudgetPlan.user_id == current_user["id"],
                        BudgetPlan.type == budget.type,
                        BudgetPlan.category == budget.category,
                        BudgetPlan.year == budget.year,
                        BudgetPlan.month.is_(None),
                    ).first()

                    if yearly_budget:
                        yearly_budget.amount = yearly_total
                    else:
                        yearly_budget = BudgetPlan(
                            user_id=current_user["id"],
                            type=budget.type,
                            category=budget.category,
                            year=budget.year,
                            month=None,
                            amount=yearly_total,
                        )
                        session.add(yearly_budget)

            session.commit()

        return BudgetPlanResponse.model_validate(existing)

    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.delete("/api/budgets/{budget_id}")
def delete_budget(budget_id: int, current_user: dict = Depends(get_current_user)):
    """Delete a budget plan."""
    session = db_manager.get_session()
    try:
        budget = session.query(BudgetPlan).filter(
            BudgetPlan.id == budget_id,
            BudgetPlan.user_id == current_user["id"]
        ).first()
        if not budget:
            raise HTTPException(status_code=404, detail="Budget not found")

        session.delete(budget)
        session.commit()
        return {"message": "Budget deleted"}
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.post("/api/budgets/bulk-delete")
def bulk_delete_budgets(budget_ids: List[int], current_user: dict = Depends(get_current_user)):
    """Delete multiple budget plans at once."""
    session = db_manager.get_session()
    try:
        deleted_count = session.query(BudgetPlan).filter(
            BudgetPlan.id.in_(budget_ids),
            BudgetPlan.user_id == current_user["id"]
        ).delete(synchronize_session=False)

        session.commit()
        return {"message": f"Deleted {deleted_count} budget(s)", "count": deleted_count}
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


# ============== Category Endpoints ==============

@app.get("/api/categories", response_model=List[CategoryInfo])
def get_categories(current_user: dict = Depends(get_current_user)):
    """Get all available categories grouped by type."""
    session = db_manager.get_session()
    try:
        # Get categories from database for current user
        categories_db = session.query(Category).filter(
            Category.user_id == current_user["id"],
            Category.is_active.is_(True)
        ).order_by(Category.display_order, Category.name).all()

        # If no categories in DB, use config defaults
        if not categories_db:
            return [
                CategoryInfo(type="Income", categories=pipeline_config.categories.income_categories),
                CategoryInfo(type="Expenses", categories=pipeline_config.categories.expense_categories),
                CategoryInfo(type="Savings", categories=pipeline_config.categories.savings_categories),
            ]

        # Group by type
        grouped = {}
        for cat in categories_db:
            if cat.type not in grouped:
                grouped[cat.type] = []
            grouped[cat.type].append(cat.name)

        return [CategoryInfo(type=t, categories=cats) for t, cats in grouped.items()]
    finally:
        session.close()


@app.get("/api/categories/all", response_model=List[CategoryResponse])
def get_all_categories(current_user: dict = Depends(get_current_user)):
    """Get all categories (including inactive) for management."""
    session = db_manager.get_session()
    try:
        categories = session.query(Category).filter(
            Category.user_id == current_user["id"]
        ).order_by(Category.type, Category.display_order, Category.name).all()
        return [CategoryResponse.model_validate(cat) for cat in categories]
    finally:
        session.close()


@app.post("/api/categories", response_model=CategoryResponse, status_code=201)
def create_category(category: CategoryCreate, current_user: dict = Depends(get_current_user)):
    """Create a new category."""
    session = db_manager.get_session()
    try:
        # Check if category already exists for this user
        existing = session.query(Category).filter(
            Category.user_id == current_user["id"],
            Category.name == category.name,
            Category.type == category.type
        ).first()

        if existing:
            raise HTTPException(
                status_code=400,
                detail=f"Category '{category.name}' already exists for type '{category.type}'"
            )

        # Create new category
        new_category = Category(
            user_id=current_user["id"],
            name=category.name,
            type=category.type,
            display_order=category.display_order,
            is_active=True
        )

        session.add(new_category)
        session.commit()
        session.refresh(new_category)

        return CategoryResponse.model_validate(new_category)
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.patch("/api/categories/{category_id}", response_model=CategoryResponse)
def update_category(
    category_id: int,
    update: CategoryUpdate,
    current_user: dict = Depends(get_current_user)
):
    """Update an existing category."""
    session = db_manager.get_session()
    try:
        category = session.query(Category).filter(
            Category.id == category_id,
            Category.user_id == current_user["id"]
        ).first()

        if not category:
            raise HTTPException(status_code=404, detail="Category not found")

        # Update fields
        if update.name is not None:
            # Check for duplicate name within user's categories
            existing = session.query(Category).filter(
                Category.user_id == current_user["id"],
                Category.name == update.name,
                Category.type == (update.type if update.type else category.type),
                Category.id != category_id
            ).first()

            if existing:
                raise HTTPException(
                    status_code=400,
                    detail=f"Category '{update.name}' already exists"
                )

            category.name = update.name

        if update.type is not None:
            category.type = update.type

        if update.is_active is not None:
            category.is_active = update.is_active

        if update.display_order is not None:
            category.display_order = update.display_order

        session.commit()
        session.refresh(category)

        return CategoryResponse.model_validate(category)
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.delete("/api/categories/{category_id}")
def delete_category(category_id: int, current_user: dict = Depends(get_current_user)):
    """Delete a category (soft delete by setting is_active=False)."""
    session = db_manager.get_session()
    try:
        category = session.query(Category).filter(
            Category.id == category_id,
            Category.user_id == current_user["id"]
        ).first()

        if not category:
            raise HTTPException(status_code=404, detail="Category not found")

        # Check if category is used in user's transactions or budgets
        transaction_count = session.query(func.count(Transaction.id)).filter(
            Transaction.user_id == current_user["id"],
            Transaction.category == category.name
        ).scalar()

        budget_count = session.query(func.count(BudgetPlan.id)).filter(
            BudgetPlan.user_id == current_user["id"],
            BudgetPlan.category == category.name
        ).scalar()

        if transaction_count > 0 or budget_count > 0:
            # Soft delete
            category.is_active = False
            session.commit()
            return {
                "message": f"Category deactivated (used in {transaction_count} transactions and {budget_count} budgets)"
            }
        else:
            # Hard delete if unused
            session.delete(category)
            session.commit()
            return {"message": "Category deleted successfully"}
    except HTTPException:
        session.rollback()
        raise
    except Exception as e:
        session.rollback()
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        session.close()


@app.get("/api/types")
def get_types():
    """Get all valid transaction types."""
    return pipeline_config.categories.valid_types


# ============== Dashboard Endpoints ==============

@app.get("/api/dashboard/summary", response_model=DashboardSummary)
def get_dashboard_summary(year: int, month: Optional[int] = None, current_user: dict = Depends(get_current_user)):
    """Get budget vs actual summary for dashboard."""
    session = db_manager.get_session()
    try:
        # Get actual spending from transactions for current user
        actual_query = session.query(
            Transaction.type,
            Transaction.category,
            func.sum(Transaction.amount).label("total"),
        ).filter(
            Transaction.user_id == current_user["id"],
            Transaction.year == year
        )

        if month:
            actual_query = actual_query.filter(Transaction.month == month)

        actual_query = actual_query.group_by(Transaction.type, Transaction.category)
        actuals = {(r.type, r.category): float(r.total) for r in actual_query.all()}

        # Get budgets - aggregate monthly budgets if viewing full year
        if month:
            # For a specific month: get that month's budget OR yearly budget (divided by 12)
            budget_query = session.query(BudgetPlan).filter(
                BudgetPlan.user_id == current_user["id"],
                BudgetPlan.year == year
            )
            budget_query = budget_query.filter(
                (BudgetPlan.month == month) | (BudgetPlan.month.is_(None))
            )

            budgets = {}
            for b in budget_query.all():
                key = (b.type, b.category)
                if b.month is None:
                    # Yearly budget - divide by 12 for monthly view
                    budgets[key] = float(b.amount) / 12
                else:
                    # Monthly budget takes precedence
                    budgets[key] = float(b.amount)
        else:
            # For full year: sum all monthly budgets OR use yearly budget
            all_budgets = session.query(BudgetPlan).filter(
                BudgetPlan.user_id == current_user["id"],
                BudgetPlan.year == year
            ).all()

            budgets = {}
            yearly_budgets = {}  # Track yearly budgets
            monthly_sums = {}  # Track sum of monthly budgets

            for b in all_budgets:
                key = (b.type, b.category)
                if b.month is None:
                    yearly_budgets[key] = float(b.amount)
                else:
                    if key not in monthly_sums:
                        monthly_sums[key] = 0.0
                    monthly_sums[key] += float(b.amount)

            # Prefer yearly budget if it exists, otherwise use monthly sum
            for key in set(list(yearly_budgets.keys()) + list(monthly_sums.keys())):
                if key in yearly_budgets:
                    # Yearly budget takes precedence
                    budgets[key] = yearly_budgets[key]
                elif key in monthly_sums:
                    # No yearly budget, sum up monthly budgets
                    budgets[key] = monthly_sums[key]

        # Build summary for each type
        def build_summary(trans_type: str, categories: List[str]) -> List[SummaryItem]:
            items = []
            # Include all categories that have either budget OR actual
            all_cats = set(categories)
            # Also include categories that have actuals even if not in the predefined list
            for type_, cat in actuals.keys():
                if type_ == trans_type:
                    all_cats.add(cat)

            for cat in all_cats:
                actual = actuals.get((trans_type, cat), 0.0)
                budget = budgets.get((trans_type, cat), 0.0)
                remaining = max(0, budget - actual)

                # Calculate percentage
                if budget > 0:
                    percent = (actual / budget * 100)
                elif actual > 0:
                    # Has actual but no budget - show as over 100%
                    percent = 100.0
                else:
                    percent = 0.0

                if actual > 0 or budget > 0:
                    items.append(SummaryItem(
                        type=trans_type,
                        category=cat,
                        budget=budget,
                        actual=actual,
                        remaining=remaining,
                        percent_complete=round(percent, 1),
                    ))
            return sorted(items, key=lambda x: x.actual, reverse=True)

        income_summary = build_summary("Income", pipeline_config.categories.income_categories)
        expense_summary = build_summary("Expenses", pipeline_config.categories.expense_categories)
        savings_summary = build_summary("Savings", pipeline_config.categories.savings_categories)

        # Calculate totals
        total_income_actual = sum(i.actual for i in income_summary)
        total_income_budget = sum(i.budget for i in income_summary)
        total_expense_actual = sum(i.actual for i in expense_summary)
        total_expense_budget = sum(i.budget for i in expense_summary)
        total_savings_actual = sum(i.actual for i in savings_summary)
        total_savings_budget = sum(i.budget for i in savings_summary)

        # Calculate Fixed Cost Ratio = (Housing + Health + Tax) / Total Income
        fixed_cost_categories = ["Housing", "Health Insurance", "Health Other", "Tax"]
        total_fixed_costs = sum(
            item.actual for item in expense_summary
            if item.category in fixed_cost_categories
        )
        fixed_cost_ratio = (total_fixed_costs / total_income_actual * 100) if total_income_actual > 0 else 0.0

        # Get previous period data for year-over-year comparison
        previous_year = year - 1
        previous_month = month  # Same month last year, or None for full year

        # Query previous period net balance
        prev_actual_query = session.query(
            Transaction.type,
            func.sum(Transaction.amount).label("total"),
        ).filter(
            Transaction.user_id == current_user["id"],
            Transaction.year == previous_year
        )

        if previous_month:
            prev_actual_query = prev_actual_query.filter(Transaction.month == previous_month)

        prev_actual_query = prev_actual_query.group_by(Transaction.type)
        prev_actuals = {r.type: float(r.total) for r in prev_actual_query.all()}

        prev_income = prev_actuals.get("Income", 0.0)
        prev_expenses = prev_actuals.get("Expenses", 0.0)
        prev_savings = prev_actuals.get("Savings", 0.0)
        prev_net = prev_income - prev_expenses - prev_savings

        return DashboardSummary(
            year=year,
            month=month,
            income=income_summary,
            expenses=expense_summary,
            savings=savings_summary,
            totals={
                "income": {"actual": total_income_actual, "budget": total_income_budget},
                "expenses": {"actual": total_expense_actual, "budget": total_expense_budget},
                "savings": {"actual": total_savings_actual, "budget": total_savings_budget},
                "net": {
                    "actual": total_income_actual - total_expense_actual - total_savings_actual,
                    "budget": total_income_budget - total_expense_budget - total_savings_budget,
                },
            },
            fixed_cost_ratio=round(fixed_cost_ratio, 1),
            previous_period={
                "year": previous_year,
                "month": previous_month,
                "net": prev_net,
            },
        )

    finally:
        session.close()


@app.get("/api/dashboard/monthly-trend")
def get_monthly_trend(year: int, categories: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    """Get monthly spending trend for the year, optionally filtered by categories (comma-separated)."""
    session = db_manager.get_session()
    try:
        # Get actual transactions
        query = session.query(
            Transaction.month,
            Transaction.type,
            func.sum(Transaction.amount).label("total"),
        ).filter(
            Transaction.user_id == current_user["id"],
            Transaction.year == year
        )

        # Add category filter if provided (comma-separated list)
        if categories:
            category_list = [c.strip() for c in categories.split(',')]
            query = query.filter(Transaction.category.in_(category_list))

        query = query.group_by(
            Transaction.month, Transaction.type
        ).order_by(Transaction.month)

        results = query.all()

        # Get budget data
        budget_query = session.query(
            BudgetPlan.month,
            BudgetPlan.type,
            func.sum(BudgetPlan.amount).label("total"),
        ).filter(
            BudgetPlan.user_id == current_user["id"],
            BudgetPlan.year == year,
            BudgetPlan.month.isnot(None)
        )

        # Add category filter for budgets if provided
        if categories:
            category_list = [c.strip() for c in categories.split(',')]
            budget_query = budget_query.filter(BudgetPlan.category.in_(category_list))

        budget_query = budget_query.group_by(
            BudgetPlan.month, BudgetPlan.type
        ).order_by(BudgetPlan.month)

        budget_results = budget_query.all()

        # Organize by month
        months = {i: {
            "month": i,
            "Income": 0,
            "Expenses": 0,
            "Savings": 0,
            "IncomeBudget": 0,
            "ExpensesBudget": 0,
            "SavingsBudget": 0
        } for i in range(1, 13)}

        # Fill in actual data
        for r in results:
            if r.type in months[r.month]:
                months[r.month][r.type] = float(r.total)

        # Fill in budget data
        for r in budget_results:
            if r.type == "Income":
                months[r.month]["IncomeBudget"] = float(r.total)
            elif r.type == "Expenses":
                months[r.month]["ExpensesBudget"] = float(r.total)
            elif r.type == "Savings":
                months[r.month]["SavingsBudget"] = float(r.total)

        return list(months.values())

    finally:
        session.close()


@app.get("/api/years")
def get_available_years(current_user: dict = Depends(get_current_user)):
    """Get list of years with transaction data."""
    session = db_manager.get_session()
    try:
        years = session.query(Transaction.year).filter(
            Transaction.user_id == current_user["id"]
        ).distinct().order_by(Transaction.year.desc()).all()
        return [y[0] for y in years]
    finally:
        session.close()


@app.get("/api/export/excel")
def export_to_excel(
    year: int,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
):
    """Export budget vs actual and categorized transactions to Excel."""
    session = db_manager.get_session()
    try:
        # Create workbook
        wb = Workbook()

        # === Sheet 1: Budget vs Actual ===
        ws_budget = wb.active
        ws_budget.title = "Budget vs Actual"

        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Title
        period_text = f"{year}" if not month else f"{['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'][month-1]} {year}"
        ws_budget['A1'] = f"Budget vs Actual - {period_text}"
        ws_budget['A1'].font = Font(bold=True, size=14)
        ws_budget.merge_cells('A1:F1')

        # Headers
        headers = ['Type', 'Category', 'Budget', 'Actual', 'Remaining', '% Complete']
        for col_num, header in enumerate(headers, 1):
            cell = ws_budget.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Get budget data for current user
        budget_query = session.query(BudgetPlan).filter(
            BudgetPlan.user_id == current_user["id"],
            BudgetPlan.year == year
        )
        if month:
            budget_query = budget_query.filter(BudgetPlan.month == month)
        else:
            budget_query = budget_query.filter(BudgetPlan.month.is_(None))
        budgets = budget_query.all()

        # Get actual transactions for current user
        trans_query = session.query(
            Transaction.type,
            Transaction.category,
            func.sum(Transaction.amount).label("total")
        ).filter(
            Transaction.user_id == current_user["id"],
            Transaction.year == year
        )

        if month:
            trans_query = trans_query.filter(Transaction.month == month)

        trans_query = trans_query.group_by(Transaction.type, Transaction.category)
        actuals = {(r.type, r.category): float(r.total) for r in trans_query.all()}

        # Build data structure
        data_by_type = {'Income': [], 'Expenses': [], 'Savings': []}

        for budget in budgets:
            actual = actuals.get((budget.type, budget.category), 0.0)
            remaining = float(budget.amount) - actual
            percent = (actual / float(budget.amount) * 100) if budget.amount > 0 else 0

            data_by_type[budget.type].append({
                'category': budget.category,
                'budget': float(budget.amount),
                'actual': actual,
                'remaining': remaining,
                'percent': percent
            })

        # Write data
        row = 4
        for trans_type in ['Income', 'Expenses', 'Savings']:
            type_data = data_by_type[trans_type]
            if not type_data:
                continue

            # Type totals
            type_budget = sum(d['budget'] for d in type_data)
            type_actual = sum(d['actual'] for d in type_data)
            type_remaining = type_budget - type_actual
            type_percent = (type_actual / type_budget * 100) if type_budget > 0 else 0

            # Type header row
            ws_budget.cell(row=row, column=1).value = trans_type
            ws_budget.cell(row=row, column=1).font = Font(bold=True)
            ws_budget.cell(row=row, column=3).value = type_budget
            ws_budget.cell(row=row, column=3).font = Font(bold=True)
            ws_budget.cell(row=row, column=3).number_format = '#,##0.00'
            ws_budget.cell(row=row, column=4).value = type_actual
            ws_budget.cell(row=row, column=4).font = Font(bold=True)
            ws_budget.cell(row=row, column=4).number_format = '#,##0.00'
            ws_budget.cell(row=row, column=5).value = type_remaining
            ws_budget.cell(row=row, column=5).font = Font(bold=True)
            ws_budget.cell(row=row, column=5).number_format = '#,##0.00'
            ws_budget.cell(row=row, column=6).value = type_percent / 100
            ws_budget.cell(row=row, column=6).font = Font(bold=True)
            ws_budget.cell(row=row, column=6).number_format = '0.0%'

            # Apply background color for type row
            type_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            for col in range(1, 7):
                ws_budget.cell(row=row, column=col).fill = type_fill
                ws_budget.cell(row=row, column=col).border = border

            row += 1

            # Category rows
            for item in type_data:
                ws_budget.cell(row=row, column=2).value = item['category']
                ws_budget.cell(row=row, column=3).value = item['budget']
                ws_budget.cell(row=row, column=3).number_format = '#,##0.00'
                ws_budget.cell(row=row, column=4).value = item['actual']
                ws_budget.cell(row=row, column=4).number_format = '#,##0.00'
                ws_budget.cell(row=row, column=5).value = item['remaining']
                ws_budget.cell(row=row, column=5).number_format = '#,##0.00'
                ws_budget.cell(row=row, column=6).value = item['percent'] / 100
                ws_budget.cell(row=row, column=6).number_format = '0.0%'

                # Color code based on performance
                if trans_type == 'Income':
                    # Green if over budget (good), red if under
                    if item['actual'] >= item['budget']:
                        ws_budget.cell(row=row, column=6).font = Font(color="008000")
                    else:
                        ws_budget.cell(row=row, column=6).font = Font(color="FF0000")
                else:
                    # Green if under budget (good), red if over
                    if item['actual'] <= item['budget']:
                        ws_budget.cell(row=row, column=6).font = Font(color="008000")
                    else:
                        ws_budget.cell(row=row, column=6).font = Font(color="FF0000")

                for col in range(1, 7):
                    ws_budget.cell(row=row, column=col).border = border

                row += 1

            row += 1  # Empty row between types

        # Adjust column widths
        ws_budget.column_dimensions['A'].width = 15
        ws_budget.column_dimensions['B'].width = 25
        ws_budget.column_dimensions['C'].width = 15
        ws_budget.column_dimensions['D'].width = 15
        ws_budget.column_dimensions['E'].width = 15
        ws_budget.column_dimensions['F'].width = 15

        # === Sheet 2: Categorized Transactions ===
        ws_trans = wb.create_sheet("Transactions")

        # Title
        ws_trans['A1'] = f"Categorized Transactions - {period_text}"
        ws_trans['A1'].font = Font(bold=True, size=14)
        ws_trans.merge_cells('A1:G1')

        # Headers
        trans_headers = ['Date', 'Type', 'Category', 'Amount', 'Description', 'Source', 'Month']
        for col_num, header in enumerate(trans_headers, 1):
            cell = ws_trans.cell(row=3, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = border

        # Get transactions for current user
        transactions_query = session.query(Transaction).filter(
            Transaction.user_id == current_user["id"],
            Transaction.year == year
        )
        if month:
            transactions_query = transactions_query.filter(Transaction.month == month)
        transactions = transactions_query.order_by(Transaction.date.desc()).all()

        # Write transaction data
        for idx, trans in enumerate(transactions, start=4):
            ws_trans.cell(row=idx, column=1).value = trans.date
            ws_trans.cell(row=idx, column=1).number_format = 'YYYY-MM-DD'
            ws_trans.cell(row=idx, column=2).value = trans.type
            ws_trans.cell(row=idx, column=3).value = trans.category
            ws_trans.cell(row=idx, column=4).value = float(trans.amount)
            ws_trans.cell(row=idx, column=4).number_format = '#,##0.00'
            ws_trans.cell(row=idx, column=5).value = trans.description or ''
            ws_trans.cell(row=idx, column=6).value = trans.source
            ws_trans.cell(row=idx, column=7).value = trans.month

            # Apply borders
            for col in range(1, 8):
                ws_trans.cell(row=idx, column=col).border = border

        # Adjust column widths
        ws_trans.column_dimensions['A'].width = 12
        ws_trans.column_dimensions['B'].width = 15
        ws_trans.column_dimensions['C'].width = 25
        ws_trans.column_dimensions['D'].width = 12
        ws_trans.column_dimensions['E'].width = 50
        ws_trans.column_dimensions['F'].width = 10
        ws_trans.column_dimensions['G'].width = 8

        # Save to BytesIO
        excel_file = BytesIO()
        wb.save(excel_file)
        excel_file.seek(0)

        # Create filename
        filename = f"LUCID_Finance_{period_text.replace(' ', '_')}.xlsx"

        return StreamingResponse(
            excel_file,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    finally:
        session.close()


# Health check
@app.get("/api/health")
def health_check():
    """Health check endpoint."""
    return {"status": "healthy", "timestamp": datetime.now().isoformat()}


if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
