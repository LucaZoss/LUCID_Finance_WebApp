"""
Export endpoints for generating Excel reports.
"""

from typing import Optional
from datetime import datetime
from io import BytesIO
from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from ..dependencies import get_db, get_current_user
from ...data_pipeline.models import Transaction, BudgetPlan

router = APIRouter(prefix="/api/export", tags=["Export"])


@router.get("/excel")
def export_to_excel(
    year: int,
    month: Optional[int] = None,
    current_user: dict = Depends(get_current_user),
    session: Session = Depends(get_db),
):
    """Export budget vs actual and categorized transactions to Excel."""
    # Create workbook
    wb = Workbook()

    # === Sheet 1: Budget vs Actual ===
    ws_budget = wb.active
    ws_budget.title = "Budget vs Actual"

    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    period_text = f"{year}" if not month else f"{['January', 'February', 'March', 'April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December'][month-1]} {year}"
    ws_budget['A1'] = f"Budget vs Actual - {period_text}"
    ws_budget['A1'].font = Font(bold=True, size=14)
    ws_budget.merge_cells('A1:F1')

    # Headers
    headers = ['Type', 'Category', 'Budget', 'Actual', 'Remaining', '% Complete']
    for col_num, header in enumerate(headers, 1):
        cell = ws_budget.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Get budget data for current user
    budget_query = session.query(BudgetPlan).filter(
        BudgetPlan.user_id == current_user["id"],
        BudgetPlan.year == year
    )
    if month:
        budget_query = budget_query.filter(BudgetPlan.month == month)
    else:
        budget_query = budget_query.filter(BudgetPlan.month.is_(None))
    budgets = budget_query.all()

    # Get actual transactions for current user
    trans_query = session.query(
        Transaction.type,
        Transaction.category,
        func.sum(Transaction.amount).label("total")
    ).filter(
        Transaction.user_id == current_user["id"],
        Transaction.year == year
    )

    if month:
        trans_query = trans_query.filter(Transaction.month == month)

    trans_query = trans_query.group_by(Transaction.type, Transaction.category)
    actuals = {(r.type, r.category): float(r.total) for r in trans_query.all()}

    # Build data structure
    data_by_type = {'Income': [], 'Expenses': [], 'Savings': []}

    for budget in budgets:
        actual = actuals.get((budget.type, budget.category), 0.0)
        remaining = float(budget.amount) - actual
        percent = (actual / float(budget.amount) * 100) if budget.amount > 0 else 0

        data_by_type[budget.type].append({
            'category': budget.category,
            'budget': float(budget.amount),
            'actual': actual,
            'remaining': remaining,
            'percent': percent
        })

    # Write data
    row = 4
    for trans_type in ['Income', 'Expenses', 'Savings']:
        type_data = data_by_type[trans_type]
        if not type_data:
            continue

        # Type totals
        type_budget = sum(d['budget'] for d in type_data)
        type_actual = sum(d['actual'] for d in type_data)
        type_remaining = type_budget - type_actual
        type_percent = (type_actual / type_budget * 100) if type_budget > 0 else 0

        # Type header row
        ws_budget.cell(row=row, column=1).value = trans_type
        ws_budget.cell(row=row, column=1).font = Font(bold=True)
        ws_budget.cell(row=row, column=3).value = type_budget
        ws_budget.cell(row=row, column=3).font = Font(bold=True)
        ws_budget.cell(row=row, column=3).number_format = '#,##0.00'
        ws_budget.cell(row=row, column=4).value = type_actual
        ws_budget.cell(row=row, column=4).font = Font(bold=True)
        ws_budget.cell(row=row, column=4).number_format = '#,##0.00'
        ws_budget.cell(row=row, column=5).value = type_remaining
        ws_budget.cell(row=row, column=5).font = Font(bold=True)
        ws_budget.cell(row=row, column=5).number_format = '#,##0.00'
        ws_budget.cell(row=row, column=6).value = type_percent / 100
        ws_budget.cell(row=row, column=6).font = Font(bold=True)
        ws_budget.cell(row=row, column=6).number_format = '0.0%'

        # Apply background color for type row
        type_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
        for col in range(1, 7):
            ws_budget.cell(row=row, column=col).fill = type_fill
            ws_budget.cell(row=row, column=col).border = border

        row += 1

        # Category rows
        for item in type_data:
            ws_budget.cell(row=row, column=2).value = item['category']
            ws_budget.cell(row=row, column=3).value = item['budget']
            ws_budget.cell(row=row, column=3).number_format = '#,##0.00'
            ws_budget.cell(row=row, column=4).value = item['actual']
            ws_budget.cell(row=row, column=4).number_format = '#,##0.00'
            ws_budget.cell(row=row, column=5).value = item['remaining']
            ws_budget.cell(row=row, column=5).number_format = '#,##0.00'
            ws_budget.cell(row=row, column=6).value = item['percent'] / 100
            ws_budget.cell(row=row, column=6).number_format = '0.0%'

            # Color code based on performance
            if trans_type == 'Income':
                # Green if over budget (good), red if under
                if item['actual'] >= item['budget']:
                    ws_budget.cell(row=row, column=6).font = Font(color="008000")
                else:
                    ws_budget.cell(row=row, column=6).font = Font(color="FF0000")
            else:
                # Green if under budget (good), red if over
                if item['actual'] <= item['budget']:
                    ws_budget.cell(row=row, column=6).font = Font(color="008000")
                else:
                    ws_budget.cell(row=row, column=6).font = Font(color="FF0000")

            for col in range(1, 7):
                ws_budget.cell(row=row, column=col).border = border

            row += 1

        row += 1  # Empty row between types

    # Adjust column widths
    ws_budget.column_dimensions['A'].width = 15
    ws_budget.column_dimensions['B'].width = 25
    ws_budget.column_dimensions['C'].width = 15
    ws_budget.column_dimensions['D'].width = 15
    ws_budget.column_dimensions['E'].width = 15
    ws_budget.column_dimensions['F'].width = 15

    # === Sheet 2: Categorized Transactions ===
    ws_trans = wb.create_sheet("Transactions")

    # Title
    ws_trans['A1'] = f"Categorized Transactions - {period_text}"
    ws_trans['A1'].font = Font(bold=True, size=14)
    ws_trans.merge_cells('A1:G1')

    # Headers
    trans_headers = ['Date', 'Type', 'Category', 'Amount', 'Description', 'Source', 'Month']
    for col_num, header in enumerate(trans_headers, 1):
        cell = ws_trans.cell(row=3, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border

    # Get transactions for current user
    transactions_query = session.query(Transaction).filter(
        Transaction.user_id == current_user["id"],
        Transaction.year == year
    )
    if month:
        transactions_query = transactions_query.filter(Transaction.month == month)
    transactions = transactions_query.order_by(Transaction.date.desc()).all()

    # Write transaction data
    for idx, trans in enumerate(transactions, start=4):
        ws_trans.cell(row=idx, column=1).value = trans.date
        ws_trans.cell(row=idx, column=1).number_format = 'YYYY-MM-DD'
        ws_trans.cell(row=idx, column=2).value = trans.type
        ws_trans.cell(row=idx, column=3).value = trans.category
        ws_trans.cell(row=idx, column=4).value = float(trans.amount)
        ws_trans.cell(row=idx, column=4).number_format = '#,##0.00'
        ws_trans.cell(row=idx, column=5).value = trans.description or ''
        ws_trans.cell(row=idx, column=6).value = trans.source
        ws_trans.cell(row=idx, column=7).value = trans.month

        # Apply borders
        for col in range(1, 8):
            ws_trans.cell(row=idx, column=col).border = border

    # Adjust column widths
    ws_trans.column_dimensions['A'].width = 12
    ws_trans.column_dimensions['B'].width = 15
    ws_trans.column_dimensions['C'].width = 25
    ws_trans.column_dimensions['D'].width = 12
    ws_trans.column_dimensions['E'].width = 50
    ws_trans.column_dimensions['F'].width = 10
    ws_trans.column_dimensions['G'].width = 8

    # Save to BytesIO
    excel_file = BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)

    # Create filename
    filename = f"LUCID_Finance_{period_text.replace(' ', '_')}.xlsx"

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
